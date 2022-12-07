
import openpyxl as op
ws=op.load_workbook("test.xlsx")
print(ws["工作表1"].max_column)

